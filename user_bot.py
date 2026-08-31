"""
User Bot - the bot regular users interact with.
"""

import io
import logging
import os
import re
import telegram
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

import database as db
from config import USER_BOT_TOKEN, ADMIN_BOT_TOKEN, ADMIN_CHAT_ID, ADMIN_USER_IDS, BKASH_NUMBER, ROCKET_NUMBER, SUPPORT_LINK, PRODUCT_PRICE

logger = logging.getLogger(__name__)

# -- Cross-bot helper ------------------------------------------

def _admin_bot():
    """Lightweight Bot instance for the Admin Bot token (used to relay submissions)."""
    return telegram.Bot(token=ADMIN_BOT_TOKEN)


# -- Credit notification (called from Admin Bot side) ---------

async def notify_user_credit_change(user_id: int, amount: int, action: str, new_balance: int):
    """Send a credit-change notification to a user via the User Bot's own token."""
    if action == "add":
        text = f"✅ {amount} credits has been added to your account by admin! New balance: {new_balance}"
    else:
        text = f"➖ {amount} credits have been deducted from your account. New balance: {new_balance}"
    try:
        async with telegram.Bot(token=USER_BOT_TOKEN) as bot:
            await bot.send_message(chat_id=user_id, text=text)
    except telegram.error.Forbidden:
        pass
    except Exception:
        pass


# -- Maintenance Mode gatekeeper -------------------------------

def _is_admin_user(user_id: int) -> bool:
    return user_id in ADMIN_USER_IDS


def _maintenance_guard(handler):
    """Wrap a user-bot handler so non-admin users are blocked during maintenance."""
    async def wrapped(update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id if update.effective_user else None
        if user_id is not None and not _is_admin_user(user_id) and db.is_maintenance():
            try:
                chat_id = (update.effective_chat.id if update.effective_chat
                           else (update.callback_query.message.chat.id if update.callback_query and update.callback_query.message else None))
                if chat_id is not None:
                    await context.bot.send_message(chat_id=chat_id, text=MAINTENANCE_TEXT, parse_mode="HTML")
            except Exception:
                pass
            return
        return await handler(update, context)
    return wrapped


# -- Keyboard --------------------------------------------------
MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        [KeyboardButton("\U0001f6cd\ufe0f Buy Products")],
        [KeyboardButton("\U0001f4b3 Deposit"), KeyboardButton("\U0001f4e5 Submit Job")],
        [KeyboardButton("\U0001f464 Profile"), KeyboardButton("\U0001f4de Support")],
    ],
    resize_keyboard=True,
)

MAINTENANCE_TEXT = (
    "<b>\U0001f6a7 System Under Maintenance</b>\n\n"
    "<b>We are currently performing routine updates. Please try again shortly!</b>"
)


def _clear_deposit_state(context: ContextTypes.DEFAULT_TYPE):
    for k in ("deposit_method", "deposit_amount", "deposit_step", "deposit_msg_id"):
        context.user_data.pop(k, None)

def _clear_product_state(context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop("awaiting_product_qty", None)
    context.user_data.pop("selected_product", None)

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # clear all pending states: deposit, submission, nid lock, product
    _clear_deposit_state(context)
    _clear_product_state(context)
    context.user_data.pop("awaiting_submission", None)
    context.user_data.pop("getting_nid", None)
    await update.message.reply_text("\u274c Cancelled. Use buttons to start again.", reply_markup=MAIN_KEYBOARD)

# -- /start ----------------------------------------------------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    db.register_user(user.id, user.username, user.first_name)
    _clear_deposit_state(context)
    _clear_product_state(context)
    context.user_data.pop("awaiting_submission", None)
    if db.is_banned(user.id):
        await update.message.reply_text("\U0001f6ab You are banned from using this bot.")
        return
    await update.message.reply_text("Welcome! Use the buttons below.", reply_markup=MAIN_KEYBOARD)


# -- Get Pic ---------------------------------------------------
# Fix future Get Pic: claim photo first, send, THEN deduct.
# If send fails, refund the claim (unmark is_sent) — no credit lost.
# Cross-bot file_id (admin bot -> user bot) fails with BadRequest → fallback: download via Admin Bot and re-upload.

async def get_pic(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _clear_deposit_state(context)
    _clear_product_state(context)
    user_id = update.effective_user.id
    if db.is_banned(user_id):
        await update.message.reply_text("\U0001f6ab You are banned from using this bot.")
        return

    # Prevent double-tap race: ignore if already processing
    if context.user_data.get("getting_nid"):
        await update.message.reply_text("\u23f3 Please wait — your previous request is processing...")
        return
    context.user_data["getting_nid"] = True

    try:
        # 2m30s cooldown after last successful Get Nid
        remaining = db.get_cooldown_remaining(user_id)
        if remaining > 0:
            mins, secs = divmod(remaining, 60)
            await update.message.reply_text(f"\u23f3 Please wait {mins}m {secs}s before next Get Nid. Cooldown: 2m30s after each successful delivery.")
            return

        credits = db.get_credits(user_id)
        if credits < 1:
            await update.message.reply_text("\u274c Insufficient credit. Contact admin to top up.")
            return

        photo_id, file_id = db.claim_photo_for_user(user_id)
        if photo_id is None:
            await update.message.reply_text("\U0001f614 No Nid's available right now. Please check back later.")
            return

        # Deduct BEFORE sending so balance can be in caption
        new_balance = db.confirm_photo_delivery(user_id)
        if new_balance is None:
            db.refund_photo_claim(photo_id)
            await update.message.reply_text("\u274c Insufficient credit. Contact admin to top up.")
            return

        caption = f"\u2705 Here's your photo! 1 credit deducted. Remaining balance: {new_balance}"

        try:
            await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="upload_photo")
        except Exception:
            pass

        # Try direct file_id first (fast path for already-converted photos)
        try:
            await update.message.reply_photo(photo=file_id, caption=caption)
            db.set_last_nid(user_id)
            # Send countdown notice
            try:
                await update.message.reply_text(f"\u23f3 Next Get Nid available in 2m 30s. Cooldown active.")
            except Exception:
                pass
            return
        except telegram.error.BadRequest as e:
            logger.warning("Get Nid direct file_id failed photo %s: %s — fallback", photo_id, e)
            # Fallback: download via whichever bot owns the file_id, then re-upload with caption
            bio = None
            for token in (ADMIN_BOT_TOKEN, USER_BOT_TOKEN):
                try:
                    async with telegram.Bot(token=token) as bot:
                        tg_file = await bot.get_file(file_id)
                        tmp = io.BytesIO()
                        await tg_file.download_to_memory(tmp)
                        tmp.seek(0)
                        tmp.name = "photo.jpg"
                        bio = tmp
                        break
                except Exception as fe:
                    logger.warning("Fallback get_file with token %s... failed: %s", token[:10], fe)
                    continue
            if bio is None:
                logger.error("Get Nid fallback get_file failed for photo %s", photo_id, exc_info=True)
                db.refund_photo_and_credit(photo_id, user_id)
                await update.message.reply_text("\u26a0\ufe0f Photo delivery failed. Your credit was refunded. Please try again — admin should re-add the photo.")
                return
            try:
                sent = await update.message.reply_photo(photo=bio, caption=caption)
                try:
                    if sent.photo:
                        db.update_photo_file_id(photo_id, sent.photo[-1].file_id)
                except Exception:
                    pass
                db.set_last_nid(user_id)
                try:
                    await update.message.reply_text(f"\u23f3 Next Get Nid available in 2m 30s. Cooldown active.")
                except Exception:
                    pass
                return
            except Exception as e2:
                logger.error("Get Nid fallback send failed photo %s: %s", photo_id, e2, exc_info=True)
                db.refund_photo_and_credit(photo_id, user_id)
                await update.message.reply_text("\u26a0\ufe0f Photo delivery failed. Your credit was refunded. Please try again.")
                return
        except Exception as e:
            logger.error("Get Nid send failed photo %s: %s", photo_id, e, exc_info=True)
            db.refund_photo_and_credit(photo_id, user_id)
            await update.message.reply_text("\u26a0\ufe0f Photo delivery failed. Your credit was refunded. Please try again.")
            return
    finally:
        context.user_data["getting_nid"] = False


# -- Submit Job ------------------------------------------------

async def submit_job(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _clear_deposit_state(context)
    _clear_product_state(context)
    if db.is_banned(update.effective_user.id):
        await update.message.reply_text("\U0001f6ab You are banned from using this bot.")
        return
    context.user_data["awaiting_submission"] = True
    await update.message.reply_text(
        "\U0001f4dd Please send your Job now.\n"
        "Submit Job Like The Format Below \U0001f447\U0001f3fb\n"
        "--------------------------------------------------------------\n"
        "\U0001f587\ufe0fAccount Link:\n"
        "\n"
        "\n"
        "\U0001f5102 Factor Code:\n"
        "\n"
        "\n"
        "\u2709\ufe0fMail:"
    )


async def handle_submission(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get("awaiting_submission"):
        return
    context.user_data["awaiting_submission"] = False

    user = update.effective_user
    message = update.message

    if message.photo:
        # Photo may have caption (user text with photo) — handle before text check
        content_type = "photo"
        content = message.photo[-1].file_id
        user_caption = message.caption or None
        await _relay_submission(user, content_type, content, user_caption)
    elif message.document:
        content_type = "document"
        content = message.document.file_id
        user_caption = message.caption or None
        await _relay_submission(user, content_type, content, user_caption)
    elif message.text:
        content_type = "text"
        content = message.text
        await _relay_submission(user, content_type, content, None)
    else:
        await message.reply_text("\u26a0\ufe0f Unsupported type. Please send text, a photo, or a document.")
        return

    await message.reply_text("\u2705 Your task has been submitted successfully!")


async def _relay_submission(user, content_type: str, content: str, user_caption: str | None = None):
    db.add_submission(user.id, content_type, content, user_caption)

    header = f"\U0001f4e5 New Job Submission\nFrom: @{user.username} (ID: {user.id})"
    try:
        async with telegram.Bot(token=ADMIN_BOT_TOKEN) as admin_bot:
            if content_type == "photo":
                full_caption = header + (f"\n\n{user_caption}" if user_caption else "")
                try:
                    await admin_bot.send_photo(chat_id=ADMIN_CHAT_ID, photo=content, caption=full_caption)
                except telegram.error.BadRequest:
                    # cross-bot file_id: download via User Bot and re-upload via Admin Bot
                    logger.warning("Relay photo cross-bot, downloading via User Bot")
                    async with telegram.Bot(token=USER_BOT_TOKEN) as ubot:
                        tg_file = await ubot.get_file(content)
                        bio = io.BytesIO()
                        await tg_file.download_to_memory(bio)
                        bio.seek(0)
                        bio.name = "submission.jpg"
                        await admin_bot.send_photo(chat_id=ADMIN_CHAT_ID, photo=bio, caption=full_caption)
            elif content_type == "document":
                full_caption = header + (f"\n\n{user_caption}" if user_caption else "")
                try:
                    await admin_bot.send_document(chat_id=ADMIN_CHAT_ID, document=content, caption=full_caption)
                except telegram.error.BadRequest:
                    logger.warning("Relay document cross-bot, downloading via User Bot")
                    async with telegram.Bot(token=USER_BOT_TOKEN) as ubot:
                        tg_file = await ubot.get_file(content)
                        bio = io.BytesIO()
                        await tg_file.download_to_memory(bio)
                        bio.seek(0)
                        bio.name = "submission.file"
                        await admin_bot.send_document(chat_id=ADMIN_CHAT_ID, document=bio, caption=full_caption)
            else:
                await admin_bot.send_message(chat_id=ADMIN_CHAT_ID, text=f"{header}\n\n{content}")
    except telegram.error.Forbidden:
        logger.warning("Relay: Admin blocked bot or wrong ADMIN_CHAT_ID")
    except Exception as e:
        logger.warning("Relay submission failed: %s", e, exc_info=True)


# -- Balance (Credits + BDT) ----------------------------------

async def balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _clear_deposit_state(context)
    _clear_product_state(context)
    user_id = update.effective_user.id
    if db.is_banned(user_id):
        await update.message.reply_text("\U0001f6ab You are banned from using this bot.")
        return
    u = db.get_user(user_id)
    credits = u["credits"] if u else 0
    bdt = db.get_bdt_balance(user_id)
    username = f"@{update.effective_user.username}" if update.effective_user.username else f"ID:{user_id}"
    await update.message.reply_text(
        f"\U0001f464 User: {username}\n"
        f"\U0001f194 ID: {user_id}\n"
        f"\U0001fa99 Current Credits: {credits}\n"
        f"\u09f3 Main Balance: {bdt} BDT"
    )

# -- Profile (user account details) ----------------------------

async def profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _clear_deposit_state(context)
    _clear_product_state(context)
    user_id = update.effective_user.id
    if db.is_banned(user_id):
        await update.message.reply_text("\U0001f6ab You are banned from using this bot.")
        return
    u = db.get_user(user_id)
    credits = u["credits"] if u else 0
    bdt = db.get_bdt_balance(user_id)
    username = f"@{update.effective_user.username}" if update.effective_user.username else f"ID:{user_id}"
    join_date = (u["joined_at"] if u and u["joined_at"] else "N/A")
    join_str = str(join_date).split("T")[0] if join_date != "N/A" else "N/A"
    await update.message.reply_text(
        f"\U0001f464 <b>User Profile</b>\n"
        f"<b>User:</b> {username}\n"
        f"<b>ID:</b> <code>{user_id}</code>\n"
        f"<b>----------------------</b>\n"
        f"\U0001f45b <b>Main Balance:</b> {bdt} BDT\n"
        f"\U0001fa99 <b>Credit Balance:</b> {credits}\n"
        f"<b>----------------------</b>\n"
        f"\U0001f4c5 <b>Joined:</b> {join_str}",
        parse_mode="HTML",
    )

# -- History ---------------------------------------------------

async def history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _clear_deposit_state(context)
    _clear_product_state(context)
    user_id = update.effective_user.id
    if db.is_banned(user_id):
        await update.message.reply_text("\U0001f6ab You are banned from using this bot.")
        return
    subs = db.get_submission_count(user_id)
    spent = db.get_user_spent(user_id)
    deposits = db.get_user_lifetime_deposits(user_id)
    await update.message.reply_text(
        f"\U0001f4c1 <b>Your Activity</b>\n"
        f"<b>----------------------</b>\n"
        f"\U0001f4dd <b>Total Submissions:</b> {subs}\n"
        f"\U0001f6cd\ufe0f <b>Total Spent:</b> {spent:.2f} BDT\n"
        f"\U0001f4b3 <b>Lifetime Deposits:</b> {deposits:.2f} BDT",
        parse_mode="HTML",
    )

# -- Products (Multi-category Inline Navigation + Confirmation) --

from config import PRODUCT_CATALOG

def _clear_pending_order(context: ContextTypes.DEFAULT_TYPE):
    for k in ("pending_order", "awaiting_product_qty", "selected_product"):
        context.user_data.pop(k, None)

async def products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _clear_deposit_state(context)
    _clear_product_state(context)
    _clear_pending_order(context)
    cats = list(PRODUCT_CATALOG.keys())
    if not cats:
        await update.message.reply_text("\u26a0\ufe0f No categories configured.")
        return
    rows = [[InlineKeyboardButton(cat, callback_data=f"prod_cat:{i}")] for i, cat in enumerate(cats)]
    rows.append([InlineKeyboardButton("\u274c Close", callback_data="prod_close")])
    await update.message.reply_text("\U0001f4c2 <b>Select Category</b>", reply_markup=InlineKeyboardMarkup(rows), parse_mode="HTML")

async def products_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    await query.answer()
    cats = list(PRODUCT_CATALOG.keys())
    if data.startswith("prod_cat:"):
        try:
            idx = int(data.split(":")[1]); cat = cats[idx]
        except Exception:
            return
        prods = PRODUCT_CATALOG.get(cat, [])
        if not prods:
            return await query.edit_message_text("\u26a0\ufe0f No products in this category.")
        rows = []
        for i, p in enumerate(prods):
            label = p['name'] if "BDT" in p['name'] else f"{p['name']} - {p['price']} BDT"
            rows.append([InlineKeyboardButton(label, callback_data=f"prod_item:{idx}:{i}")])
        rows.append([InlineKeyboardButton("\u2b05 Back", callback_data="prod_back")])
        await query.edit_message_text(f"<b>{cat}</b>\n\nPlease select a product below:", parse_mode="HTML", reply_markup=InlineKeyboardMarkup(rows))
        return
    if data == "prod_back":
        cats = list(PRODUCT_CATALOG.keys())
        rows = [[InlineKeyboardButton(cat, callback_data=f"prod_cat:{i}")] for i, cat in enumerate(cats)]
        rows.append([InlineKeyboardButton("\u274c Close", callback_data="prod_close")])
        return await query.edit_message_text("\U0001f4c2 <b>Select Category</b>", reply_markup=InlineKeyboardMarkup(rows), parse_mode="HTML")
    if data == "prod_close":
        try: await query.delete_message()
        except: await query.edit_message_text("\u274c Closed.")
        return
    if data.startswith("prod_item:"):
        try:
            _, c_idx, p_idx = data.split(":"); c_idx=int(c_idx); p_idx=int(p_idx)
            cat = cats[c_idx]; prod = PRODUCT_CATALOG[cat][p_idx]
        except Exception:
            return
        # Strip duplicate price from display name for quantity step: keep raw short name only
        raw_name = prod["name"]
        # If name already contains price (e.g. "Premium Outlook - 1.50 BDT"), use it as-is for first line
        # Do NOT prepend extra emoji — keep exact short name per GLOBAL NAMING RULE
        context.user_data["selected_product"] = {"cat": cat, "idx": p_idx, "name": raw_name, "sheet": prod.get("sheet") or prod.get("sheet_tab") or prod.get("sheet_name") or "Trusted Income Bot", "price": float(prod["price"])}
        context.user_data["awaiting_product_qty"] = True
        quantity_text = (
            f"<b>\U0001f4e7 Premium Outlook</b>\n"
            f"\U0001f4b0 <b>{prod['price']} BDT / Unit</b>\n\n"
            f"\U0001f522 <b>Enter Quantity:</b>"
        )
        await query.edit_message_text(
            quantity_text,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("\u274c Cancel", callback_data="prod_cancel")]])
        )
        return
    if data == "prod_cancel":
        _clear_pending_order(context)
        _clear_product_state(context)
        try: await query.edit_message_text("\u274c <b>Order Has Been Canceled.</b>", parse_mode="HTML")
        except: pass
        return
    if data == "prod_confirm":
        # Execute order
        order = context.user_data.get("pending_order")
        if not order:
            return await query.edit_message_text("\u26a0\ufe0f No pending order. Please start again via \U0001f6cd\ufe0f Products.")
        await query.edit_message_text("\u23f3 Processing order...")
        # Re-validate balance and stock then deduct + allocate
        user = update.effective_user
        qty = order["qty"]; price = order["price"]; total = order["total"]; sheet = order["sheet"]; prod_name = order["name"]
        bal = db.get_bdt_balance(user.id)
        if bal < total - 1e-9:
            _clear_pending_order(context)
            return await query.message.reply_text(f"\u26a0\ufe0f Insufficient BDT Balance! Please deposit first. Required: {total} BDT, Balance: {bal} BDT")
        try:
            import sheets
            avail = sheets.count_available(sheet)
        except Exception as e:
            logger.error("Sheets count failed %s: %s", sheet, e, exc_info=True)
            _clear_pending_order(context)
            return await query.message.reply_text("\u26a0\ufe0f Stock check failed. Try later.")
        if avail < qty:
            _clear_pending_order(context)
            return await query.message.reply_text(f"\u26a0\ufe0f Low Stock! Only {avail} items left for {prod_name}.")
        new_bal = db.deduct_bdt_for_purchase(user.id, total)
        if new_bal is None:
            _clear_pending_order(context)
            return await query.message.reply_text(f"\u26a0\ufe0f Insufficient BDT Balance! Required: {total} BDT")
        username = user.username or str(user.id)
        try:
            import sheets
            items = await sheets.allocate_items(username, qty, sheet)
        except ValueError as ve:
            db.refund_bdt_purchase(user.id, total)
            _clear_pending_order(context)
            return await query.message.reply_text(f"\u26a0\ufe0f {ve}")
        except Exception as e:
            db.refund_bdt_purchase(user.id, total)
            logger.error("Sheets allocate failed %s: %s", sheet, e, exc_info=True)
            _clear_pending_order(context)
            return await query.message.reply_text("\u26a0\ufe0f Allocation failed. Refunded.")
        # Excel
        import tempfile, os
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from datetime import datetime, timezone
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Order"
            ws["A1"] = "Sl No."; ws["B1"] = "Product Data / Key"
            for cell in ws[1]: cell.font = Font(bold=True)
            for i, it in enumerate(items, start=1):
                ws.cell(row=i+1, column=1, value=i); ws.cell(row=i+1, column=2, value=it)
            ws.column_dimensions["A"].width = 10; ws.column_dimensions["B"].width = 50
            ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            safe_user = (username.lstrip("@") or str(user.id)).replace("/", "_")[:20]
            fname = f"Order_{safe_user}_{ts}.xlsx"
            tmp_path = os.path.join(tempfile.gettempdir(), fname)
            wb.save(tmp_path)
            purchased_product_name = prod_name
            # Derived short name: strip trailing price suffix (e.g. "- 1.50 BDT")
            short_name = re.sub(r"\s*-\s*[\d.]+ BDT\s*$", "", purchased_product_name).strip()
            success_msg = (
                f"\u2705 <b>Mail Delivered!</b>\n\n"
                f"\U0001f4e7 <b>{qty}x {short_name}</b>\n"
                f"\U0001f4b0 <b>Paid  : {total:.2f} BDT</b>\n"
                f"\u2501━━━━━━━━━━━━━━━━━━━━\n"
                f"\U0001f4c2 <b>File below \u2193</b>"
            )
            await query.message.reply_text(success_msg, parse_mode="HTML")
            await context.bot.send_document(chat_id=update.effective_chat.id, document=open(tmp_path, "rb"), filename=fname)
            try: os.remove(tmp_path)
            except: pass
            # --- SEND NOTIFICATION TO LOG GROUP ---
            log_group_id = os.getenv("LOG_GROUP_ID")
            if log_group_id:
                try:
                    group_msg = (
                        f"🎉 <b>New Order Placed!</b>\n\n"
                        f"👤 <b>User:</b> @{update.effective_user.username or update.effective_user.first_name} [<code>{update.effective_user.id}</code>]\n"
                        f"⚡️ <b>Product:</b> 📧 Premium Outlook\n"
                        f"🪡 <b>Quantity:</b> {qty}\n"
                        f"💰 <b>Total:</b> {total:.2f} BDT\n"
                        f"👛 <b>Remaining Bal:</b> {new_bal:.2f} BDT"
                    )
                    await context.bot.send_message(chat_id=log_group_id, text=group_msg, parse_mode="HTML")
                except Exception as e:
                    print(f"Failed to send log to group: {e}")
            _clear_pending_order(context)
        except Exception as e:
            logger.error("Excel/send failed: %s", e, exc_info=True)
            await query.message.reply_text("\u26a0\ufe0f Allocated but file failed. Contact support.")
            _clear_pending_order(context)
        return

async def handle_product_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get("awaiting_product_qty"):
        return False
    text = (update.message.text or "").strip()
    if text.lower() in ("/cancel", "cancel"):
        _clear_pending_order(context)
        await update.message.reply_text("\u274c <b>Order Has Been Canceled.</b>", parse_mode="HTML")
        return True
    if not text.isdigit():
        await update.message.reply_text("\u26a0\ufe0f Please enter a valid number (e.g., 5). Send /cancel to abort.")
        return True
    qty = int(text)
    if qty <= 0:
        await update.message.reply_text("\u26a0\ufe0f Quantity must be > 0.")
        return True
    if qty > 1000:
        await update.message.reply_text("\u26a0\ufe0f Max 1000 per order.")
        return True
    sel = context.user_data.get("selected_product")
    if not sel:
        _clear_pending_order(context)
        await update.message.reply_text("\u26a0\ufe0f No product selected. Open \U0001f6cd\ufe0f Products again.")
        return True
    context.user_data.pop("awaiting_product_qty", None)
    user = update.effective_user
    total = round(qty * float(sel["price"]), 2)
    bal = db.get_bdt_balance(user.id)
    context.user_data["pending_order"] = {"name": sel["name"], "sheet": sel["sheet"], "price": float(sel["price"]), "qty": qty, "total": total}
    # Fixed format: Order Summary + 1 standard line then details, using short_name
    summary = (
        f"\U0001f4e9 <b>Order Summary</b>\n\n"
        f"\u26a1 <b>Product   : \U0001f4e7 Premium Outlook</b>\n"
        f"\U0001faa1 <b>Quantity  : {qty}</b>\n"
        f"\U0001f4b0 <b>Total     : {total:.2f} BDT</b>\n"
        f"\U0001f45b <b>Balance   : {bal:.2f} BDT</b>"
    )
    await update.message.reply_text(summary, parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("\u2705 Confirm Order", callback_data="prod_confirm"), InlineKeyboardButton("\u274c Cancel", callback_data="prod_cancel")]]))
    return True

async def support(update: Update, context: ContextTypes.DEFAULT_TYPE):
    _clear_deposit_state(context)
    _clear_product_state(context)
    await update.message.reply_text(f"\U0001f4de Support: {SUPPORT_LINK}\nContact admin for help.")

# -- Deposit Flow (bKash & Rocket ONLY - Clean State Machine) --

GATEWAY_META = {
    "bkash": {"name": "bKash", "icon": "\U0001f4b8", "number": BKASH_NUMBER},
    "rocket": {"name": "Rocket", "icon": "\U0001f680", "number": ROCKET_NUMBER},
}

async def deposit_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """STEP 1: Show payment method selection."""
    if db.is_banned(update.effective_user.id):
        await update.message.reply_text("\U0001f6ab You are banned from using this bot.")
        return
    _clear_deposit_state(context)
    _clear_product_state(context)
    context.user_data.pop("awaiting_submission", None)
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("\U0001f4b8 bKash", callback_data="dep_method:bkash"),
         InlineKeyboardButton("\U0001f680 Rocket", callback_data="dep_method:rocket")],
    ])
    await update.message.reply_text("<b>\U0001f4b0 Deposit Funds</b>\n\n<b>Choose your preferred gateway:</b>", reply_markup=keyboard, parse_mode="HTML")

async def deposit_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data

    # STEP 1 click -> STEP 2 (edit to amount prompt)
    if data.startswith("dep_method:"):
        await query.answer()
        method = data.split(":")[1]
        meta = GATEWAY_META[method]
        context.user_data["deposit_method"] = method
        context.user_data["deposit_step"] = "WAITING_FOR_AMOUNT"
        text = (
            f"<b>{meta['icon']} {meta['name']} Payment</b>\n\n"
            f"<b>Enter Deposit Amount in BDT:</b>\n"
            f"<i>(Minimum Deposit: 20 BDT)</i>"
        )
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("\u274c Cancel", callback_data="dep_cancel")]
        ]), parse_mode="HTML")
        return

    # STEP 3 click -> STEP 4 (edit to TrxID prompt)
    if data == "dep_paid":
        await query.answer()
        amt = context.user_data.get("deposit_amount")
        if not amt:
            await query.answer("Please enter amount first.", show_alert=True)
            return
        method = context.user_data.get("deposit_method")
        meta = GATEWAY_META.get(method, GATEWAY_META["bkash"])
        context.user_data["deposit_step"] = "WAITING_FOR_TRXID"
        text = (
            f"<b>\U0001f4cd Submit Transaction Key</b>\n\n"
            f"<b>Target Amount: {amt:.2f} BDT via {meta['name']}</b>\n\n"
            f"<b>Enter your SMS Transaction ID below:</b>\n"
            f"<i>(Example format: 9A8B7C6D5E)</i>"
        )
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("\u274c Cancel", callback_data="dep_cancel")]
        ]), parse_mode="HTML")
        return

    # STEP 6: Cancel at any stage
    if data == "dep_cancel":
        await query.answer()
        _clear_deposit_state(context)
        await query.edit_message_text("<b>\U0001f6ab Deposit request terminated.</b>", parse_mode="HTML")
        return

async def handle_deposit_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    step = context.user_data.get("deposit_step")
    if not step:
        return False
    text = (update.message.text or "").strip()
    user = update.effective_user

    # STEP 2: User entered amount -> STEP 3 (send NEW payment instruction message)
    if step == "WAITING_FOR_AMOUNT":
        if not text.replace(".", "", 1).isdigit():
            await update.message.reply_text("\u26a0\ufe0f Please enter a valid numeric amount (e.g., 20).")
            return True
        amount = float(text)
        if amount < 20.0:
            await update.message.reply_text("<b>\u26a0\ufe0f Minimum deposit is 20.0 BDT.</b>", parse_mode="HTML")
            return True
        context.user_data["deposit_amount"] = amount
        method = context.user_data.get("deposit_method")
        meta = GATEWAY_META.get(method, GATEWAY_META["bkash"])
        context.user_data["deposit_step"] = "WAITING_FOR_PAID_CLICK"
        text = (
            f"<b>\U0001f4e5 Transfer Instructions ({meta['name']})</b>\n\n"
            f"<b>Please Send Money {amount:.2f} BDT to:</b>\n"
            f"<code>{meta['number']}</code>\n\n"
            f"<i>(Tap the number to copy)</i>\n\n"
            f"<b>Click below after completing payment:</b>"
        )
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("\u2705 Paid", callback_data="dep_paid"),
             InlineKeyboardButton("\u274c Cancel", callback_data="dep_cancel")]
        ]), parse_mode="HTML")
        return True

    # STEP 5: User entered TrxID -> verify against pending_deposits
    if step == "WAITING_FOR_TRXID":
        trx = text.strip().upper()
        if len(trx) < 4:
            await update.message.reply_text("\u26a0\ufe0f TrxID too short. Please send valid TrxID.")
            return True
        pending = db.get_pending_deposit_by_trx(trx)
        if not pending or pending.get("status") != "UNCLAIMED":
            await update.message.reply_text("<b>\u26a0\ufe0f Invalid or already used TrxID. Please check your SMS and try again.</b>", parse_mode="HTML")
            return True
        selected = "bKash" if context.user_data.get("deposit_method") == "bkash" else "Rocket"
        actual = pending.get("gateway", "")
        if actual.lower() != selected.lower():
            await update.message.reply_text(f"<b>\u26a0\ufe0f Gateway mismatch! This TrxID belongs to {actual}, but you selected {selected}.</b>", parse_mode="HTML")
            return True
        claimed = db.claim_pending_deposit(trx, user.id)
        if claimed:
            amt = float(claimed["amount"])
            new_bal = db.get_bdt_balance(user.id)
            await update.message.reply_text(
                f"<b>\U0001f389 Payment Confirmed! Added {amt:.2f} BDT to your balance.\nCurrent Balance: {new_bal:.2f} BDT</b>",
                parse_mode="HTML"
            )
            _clear_deposit_state(context)
            return True
        await update.message.reply_text("<b>\u26a0\ufe0f Invalid or already used TrxID. Please check your SMS and try again.</b>", parse_mode="HTML")
        return True
    return False


# -- Catch-all for submissions + deposit + products -----------

async def catch_submission(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("awaiting_product_qty"):
        handled = await handle_product_quantity(update, context)
        if handled:
            return
    # deposit flow has priority over submissions
    if context.user_data.get("deposit_step"):
        handled = await handle_deposit_text(update, context)
        if handled:
            return
    await handle_submission(update, context)


# -- Build application -----------------------------------------

def build_user_bot() -> Application:
    app = Application.builder().token(USER_BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", _maintenance_guard(start)))
    app.add_handler(CommandHandler("cancel", _maintenance_guard(cancel)))
    app.add_handler(MessageHandler(filters.Regex("^\U0001f6cd\ufe0f Buy Products$"), _maintenance_guard(products)))
    app.add_handler(MessageHandler(filters.Regex("^\U0001f4b3 Deposit$"), _maintenance_guard(deposit_entry)))
    app.add_handler(MessageHandler(filters.Regex("^\U0001f4e5 Submit Job$"), _maintenance_guard(submit_job)))
    app.add_handler(MessageHandler(filters.Regex("^\U0001f464 Profile$"), _maintenance_guard(profile)))
    app.add_handler(MessageHandler(filters.Regex("^\U0001f4de Support$"), _maintenance_guard(support)))
    app.add_handler(CallbackQueryHandler(_maintenance_guard(deposit_callback), pattern=r"^dep_"))
    app.add_handler(CallbackQueryHandler(_maintenance_guard(products_callback), pattern=r"^prod_"))

    # Catch text/photo/document that arrive while awaiting a submission/deposit
    app.add_handler(MessageHandler(
        (filters.TEXT | filters.PHOTO | filters.Document.ALL) & ~filters.COMMAND,
        _maintenance_guard(catch_submission),
    ))

    return app
